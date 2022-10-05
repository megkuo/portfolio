# Megan Kuo, mlk6une
# 9.8.22
# Adopted from Kal Buterbaugh's (Class of 2022) lab_grades.py

# to handle reading from the lab attendance excel sheet, download openpyxl!
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# from __future__ import print_function
import os.path
from sys import argv

"""
----------WHAT YOU NEED TO USE THIS SCRIPT--------------------
- download the lab attendance spreadsheet that TA's use to input attendance for lab, ex. CS1110_S22_Lab_Attendance.xlsx
- a "sample_grade_file.csv" that contains the formatting Collab excel sheets use
    - edit this to change "EXACT ASSIGNMENT NAME" to the exact assignment name as listed in collab (change the variable below)

----------HOW TO USE--------------------
Using the terminal, input command should be similar to:

python .\lab_spreadsheet_to_collab.py 2 CS1110_F22_Lab_Attendance.xlsx

argv[1] = the lab number this is, ex. Lab 02 would be -> 2
argv[2] = the name of the .csv the grades are from, format is expected to be:
    Name (Last, First); Computing ID ; First lab ; Second lab ; etc.

----------WHAT TO EXPECT--------------------

Output should be a folder called "Lab __", which will have a grades.csv inside of it that can be uploaded to Collab

----------OTHER NOTES-----------------------

Limitations: 
    - uploading to Collab has not been tested at this time


"""

# Indices in data
NAME = 0 # takes up two indices when the lines are split later on, as name gets broken up into last, first
NET_ID = 1
FIRST_GRADE = 2 # May change when we add in partners

EXACT_ASSIGNMENT_NAME = "Lab01 Installing" # replace this string with the exact name of the assignment as listed in Collab

SPREADSHEET = argv[2].replace("_", " ")

GRADES = { # levels of grades
    'P': 1,
    'M': 1,
    'L': 0.7,
    'A': 0,
    'E': 1
}

LAB_SECTS = ['101', '102', '103', '104', '105', '106', '107', '108', '109']

LAB_NUM_OFFSET = 1 # index of first lab column

# dictionaries
grades = {}  # dictionary of {comp_ID : grade}
names = {}  # dictionary of {comp_ID : full_name}

lab_num = 0

# what is the number of the current lab, passed in as argument
try:
    lab_num = int(argv[1])
except (IndexError, TypeError) as e:
    print('Error reading lab number.')
    quit()


def get_labnum(num):
    """
    Turns lab number input into the correct format (ex. 2 -> 02)
    """
    if num < 10:
        return '0' + str(num)
    return str(num)

file_name = str(argv[2])
# particular file, ex. "CS1110_S22_Lab_Attendance.xlsx"

# reading through the excel sheet downloaded
all_students = [] # list of all rows of sheets
wb = load_workbook(filename = file_name)
# print(wb.sheetnames)

# print(ws.rows)

for sheet in wb:
    # print(sheet.title)
    # if sheet.title == "Lab 102":
    #         break
    if sheet.title != "Intro Sheet":
        for row in sheet.iter_rows(min_row=2, max_col=14, max_row=100, values_only=True):
            # for cell in row:
            #     print(cell)
            # print(row)
            if row[0] is not None:
                all_students.append(tuple(row))
            else:
                if row[1] is not None:
                    print("WARNING: Student was skipped-- no name associated. Computing ID: " + row[1])

# print(all_students)

# next steps:
"""
    make sure all sheets can be read from
    sort the all students list to match the one from collab
    adjust code below to iterate through all_students instead of data
"""

# with open(file_name, 'r')  as data:
#     data.readline()  # remove header row
def sort_key(tup):
    """
    lets you sort by tuples despite None values, sorts by last name
    :param tup: the tuple
    :return:
    """
    return tup[0]

all_students.sort(key=sort_key)
for student in all_students:

    # student = student.replace('"', '').replace(" ", "") # remove '"' and spaces

    # student = student.strip().split(',') # convert from string to list
    # print(student)

    if student[0] is not None:

        if lab_num >= LAB_NUM_OFFSET:
            # print("grade:", student[4])
            try:  # tries to find their grade for the lab
                # print(FIRST_GRADE, LAB_NUM_OFFSET)
                # print("Index:", str(FIRST_GRADE + (lab_num)))
                # print(student[FIRST_GRADE + (lab_num)])
                grades[student[NET_ID]] = GRADES[student[FIRST_GRADE + (lab_num - LAB_NUM_OFFSET)].upper()]
                # gets the student's grade from the proper column, turns it into a letter, then to number grade

            except (AttributeError, KeyError, IndexError) as e:  # if the student doesn't currently have a grade for it
                grades[student[NET_ID]] = 0  # defaults to 0
                # print("ERROR")
            names[student[NET_ID]] = student[NAME] # + ", " + student[NAME + 1]
        else:
            # Ungraded lab
            grades[student[NET_ID]] = 1
    else:
        if student[1] is not None:
            print("WARNING: Student was skipped-- no name associated. Computing ID: " + student[1])


# print(grades)

# generates new .csv file for the grades, takes the formatting required
with open('sample_grade_file.csv', 'r') as sample_file:
    text = sample_file.read().replace('{ASSIGNMENT_NAME}', 'Lab' + get_labnum(lab_num)).replace("EXACT ASSIGNMENT NAME", EXACT_ASSIGNMENT_NAME)

# makes folder for the lab assignment to be in
if not os.path.exists('Lab' + get_labnum(lab_num)):
    os.mkdir('Lab' + get_labnum(lab_num))

# will create new file called grades.csv
with open('Lab' + get_labnum(lab_num) + '/grades.csv', 'w') as grade_file:
    for student, grade in grades.items():
        # print(names)
        text += student + ',' + student + ',' + names[student].replace(' ', '') + ',' + str(grade)+',,\n'
        # print(str(grade))
    text = text[:-1]    # Remove extra newline
    grade_file.write(text)
